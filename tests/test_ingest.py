from __future__ import annotations

import sqlite3
from pathlib import Path

import pytest

from dbaide.ingest import ImportManifest, import_workbooks


def _q(table: str) -> str:
    return '"' + table.replace('"', '""') + '"'


def test_csv_infers_types_and_preserves_leading_zero_codes(tmp_path):
    csv = tmp_path / "sales.csv"
    csv.write_text(
        "code,amount,qty,city\n0012,1234.5,3,BJ\n0034,99,10,SH\n0056,,2,GZ\n",
        encoding="utf-8",
    )
    res = import_workbooks([csv], dest_dir=tmp_path / "imports")
    sheet = res.manifest.workbooks[0].sheets[0]
    types = {c.name: c.type for c in sheet.columns}
    assert types == {"code": "TEXT", "amount": "REAL", "qty": "INTEGER", "city": "TEXT"}

    con = sqlite3.connect(res.db_path)
    try:
        assert [r[0] for r in con.execute(f"SELECT code FROM {_q(sheet.table)}")] == ["0012", "0034", "0056"]
        assert con.execute(f"SELECT SUM(amount) FROM {_q(sheet.table)}").fetchone()[0] == 1333.5
        # blank cell became NULL, not "" or 0
        assert con.execute(f"SELECT amount FROM {_q(sheet.table)} WHERE code='0056'").fetchone()[0] is None
    finally:
        con.close()


def test_csv_sanitizes_and_dedups_columns(tmp_path):
    csv = tmp_path / "weird.csv"
    # punctuation + CJK + duplicate header + an empty header
    csv.write_text("销售额(元),销售额(元),,name\n1,2,3,x\n", encoding="utf-8")
    res = import_workbooks([csv], dest_dir=tmp_path / "imports")
    names = [c.name for c in res.manifest.workbooks[0].sheets[0].columns]
    assert names[0] == "销售额_元"        # punctuation -> _, CJK kept
    assert names[1] == "销售额_元_2"      # duplicate disambiguated
    assert names[2] == "col_3"           # empty header -> positional name
    assert names[3] == "name"


def test_unsupported_type_and_empty_file_raise(tmp_path):
    bad = tmp_path / "x.json"
    bad.write_text("{}", encoding="utf-8")
    with pytest.raises(ValueError):
        import_workbooks([bad], dest_dir=tmp_path / "i1")

    empty = tmp_path / "empty.csv"
    empty.write_text("\n\n", encoding="utf-8")
    with pytest.raises(ValueError):
        import_workbooks([empty], dest_dir=tmp_path / "i2")


def test_manifest_round_trips_and_has_provenance(tmp_path):
    csv = tmp_path / "t.csv"
    csv.write_text("a,b\n1,2\n", encoding="utf-8")
    res = import_workbooks([csv], dest_dir=tmp_path / "imports")
    loaded = ImportManifest.load(res.db_path.parent / "manifest.json")
    wb = loaded.workbooks[0]
    assert wb.source_filename == "t.csv"
    assert wb.file_hash and wb.imported_at
    assert wb.sheets[0].row_count == 1


def test_failed_import_leaves_no_partial_db(tmp_path, monkeypatch):
    csv = tmp_path / "t.csv"
    csv.write_text("a,b\n1,2\n", encoding="utf-8")
    import dbaide.ingest.importer as imp

    monkeypatch.setattr(imp, "_write_table", lambda *a, **k: (_ for _ in ()).throw(RuntimeError("boom")))
    with pytest.raises(RuntimeError):
        import_workbooks([csv], dest_dir=tmp_path / "imports")
    assert not (tmp_path / "imports" / "data.db").exists()


def test_xlsx_multi_sheet_skips_hidden_and_handles_dates(tmp_path):
    pytest.importorskip("openpyxl")
    from openpyxl import Workbook
    import datetime

    wb = Workbook()
    s1 = wb.active
    s1.title = "data"
    s1.append(["product", "price", "qty", "day"])
    s1.append(["A", 12.5, 3, datetime.date(2026, 1, 5)])
    s2 = wb.create_sheet("store")
    s2.append(["name"]); s2.append(["flagship"])
    hidden = wb.create_sheet("secret"); hidden.sheet_state = "hidden"
    hidden.append(["x"]); hidden.append([1])
    path = tmp_path / "book.xlsx"
    wb.save(path)

    res = import_workbooks([path], dest_dir=tmp_path / "imports")
    sheets = res.manifest.workbooks[0].sheets
    assert [s.sheet_name for s in sheets] == ["data", "store"]   # hidden sheet skipped
    assert sheets[0].table == "book__data"                       # named sheet → namespaced
    cols = {c.name: c.type for c in sheets[0].columns}
    assert cols == {"product": "TEXT", "price": "REAL", "qty": "INTEGER", "day": "TEXT"}
    con = sqlite3.connect(res.db_path)
    try:
        assert con.execute(f'SELECT day FROM {_q(sheets[0].table)}').fetchone()[0].startswith("2026-01-05")
    finally:
        con.close()


def test_table_name_collapses_single_csv_but_namespaces_named_sheets(tmp_path):
    a = tmp_path / "a.csv"; a.write_text("x\n1\n", encoding="utf-8")
    b = tmp_path / "b.csv"; b.write_text("y\n2\n", encoding="utf-8")
    res = import_workbooks([a, b], dest_dir=tmp_path / "imports")
    tables = sorted(s.table for w in res.manifest.workbooks for s in w.sheets)
    assert tables == ["a", "b"]   # CSV sheet == file stem → no redundant prefix


def test_collection_add_and_remove_workbooks(tmp_path):
    from dbaide.ingest import ExcelCollection

    sales = tmp_path / "sales.csv"; sales.write_text("amt\n10\n20\n", encoding="utf-8")
    cust = tmp_path / "customers.csv"; cust.write_text("name\nAda\n", encoding="utf-8")

    col = ExcelCollection(tmp_path / "imports" / "shop")
    assert not col.exists()

    col.add([sales])                       # create
    col.add([cust])                        # append a second workbook
    assert col.exists()
    books = col.workbooks()
    assert len(books) == 2
    assert {b.source_filename for b in books} == {"sales.csv", "customers.csv"}

    con = sqlite3.connect(col.db_path)
    try:
        assert {r[0] for r in con.execute("SELECT name FROM sqlite_master WHERE type='table'")} == {"sales", "customers"}
    finally:
        con.close()

    # remove the sales workbook → its table is dropped, the other survives
    sales_id = next(b.id for b in books if b.source_filename == "sales.csv")
    col.remove(sales_id)
    assert [b.source_filename for b in col.workbooks()] == ["customers.csv"]
    con = sqlite3.connect(col.db_path)
    try:
        tables = {r[0] for r in con.execute("SELECT name FROM sqlite_master WHERE type='table'")}
        assert tables == {"customers"}
    finally:
        con.close()

    with pytest.raises(KeyError):
        col.remove("wb_does_not_exist")


def test_collection_add_keeps_table_names_unique(tmp_path):
    from dbaide.ingest import ExcelCollection

    # two files that would both want the table name "data"
    f1 = tmp_path / "data.csv"; f1.write_text("a\n1\n", encoding="utf-8")
    sub = tmp_path / "sub"; sub.mkdir()
    f2 = sub / "data.csv"; f2.write_text("b\n2\n", encoding="utf-8")

    col = ExcelCollection(tmp_path / "imports" / "c")
    col.add([f1])
    col.add([f2])
    tables = sorted(s.table for w in col.workbooks() for s in w.sheets)
    assert tables == ["data", "data_2"]


def test_import_spec_logical_name_drives_table_name(tmp_path):
    from dbaide.ingest import ImportSpec

    raw = tmp_path / "客户数据导出_最终版_v2.csv"
    raw.write_text("name\nAda\n", encoding="utf-8")
    res = import_workbooks([ImportSpec(raw, name="customers")], dest_dir=tmp_path / "imports")
    sheet = res.manifest.workbooks[0].sheets[0]
    assert res.manifest.workbooks[0].name == "customers"
    assert sheet.table == "customers"
    assert sheet.display_name == "customers"


def test_collection_rename_workbook_renames_table(tmp_path):
    from dbaide.ingest import ExcelCollection

    f = tmp_path / "messy.csv"; f.write_text("v\n1\n2\n", encoding="utf-8")
    col = ExcelCollection(tmp_path / "imports" / "c")
    col.add([f])
    wid = col.workbooks()[0].id

    col.rename(wid, "sales")
    wb = col.workbooks()[0]
    assert wb.name == "sales"
    assert wb.sheets[0].table == "sales"
    con = sqlite3.connect(col.db_path)
    try:
        tables = {r[0] for r in con.execute("SELECT name FROM sqlite_master WHERE type='table'")}
        assert tables == {"sales"}
        assert con.execute("SELECT SUM(v) FROM sales").fetchone()[0] == 3
    finally:
        con.close()

    with pytest.raises(ValueError):
        col.rename(wid, "   ")


def test_collection_overwrite_add_replaces_same_name(tmp_path):
    from dbaide.ingest import ExcelCollection, ImportSpec

    v1 = tmp_path / "v1.csv"; v1.write_text("amt\n1\n2\n3\n", encoding="utf-8")
    v2 = tmp_path / "v2.csv"; v2.write_text("amt\n10\n", encoding="utf-8")
    col = ExcelCollection(tmp_path / "imports" / "c")
    col.add([ImportSpec(v1, name="sales")])
    assert col.workbooks()[0].sheets[0].row_count == 3

    # overwrite-add: same logical name → quick delete-then-add, no duplicate table
    col.add([ImportSpec(v2, name="sales")], overwrite=True)
    books = col.workbooks()
    assert len(books) == 1
    assert books[0].sheets[0].row_count == 1
    con = sqlite3.connect(col.db_path)
    try:
        tables = {r[0] for r in con.execute("SELECT name FROM sqlite_master WHERE type='table'")}
        assert tables == {"sales"}            # not sales + sales_2
    finally:
        con.close()


def test_csv_preserves_newlines_inside_quoted_fields(tmp_path):
    csv = tmp_path / "notes.csv"
    csv.write_text('id,note\n1,"line a\nline b"\n2,plain\n', encoding="utf-8")
    res = import_workbooks([csv], dest_dir=tmp_path / "imports")
    sheet = res.manifest.workbooks[0].sheets[0]
    con = sqlite3.connect(res.db_path)
    try:
        rows = con.execute(f"SELECT id, note FROM {_q(sheet.table)} ORDER BY id").fetchall()
    finally:
        con.close()
    assert rows == [(1, "line a\nline b"), (2, "plain")]   # embedded newline kept, not corrupted


def test_oversized_integers_kept_as_exact_text(tmp_path):
    csv = tmp_path / "ids.csv"
    csv.write_text("acct\n99999999999999999999\n12345678901234567890\n", encoding="utf-8")
    res = import_workbooks([csv], dest_dir=tmp_path / "imports")   # must not raise OverflowError
    sheet = res.manifest.workbooks[0].sheets[0]
    assert sheet.columns[0].type == "TEXT"
    con = sqlite3.connect(res.db_path)
    try:
        vals = [r[0] for r in con.execute(f"SELECT acct FROM {_q(sheet.table)}")]
    finally:
        con.close()
    assert vals == ["99999999999999999999", "12345678901234567890"]   # exact, not lossy REAL


def test_invalid_collection_name_rejected():
    from dbaide.ingest import is_valid_collection_name

    assert is_valid_collection_name("sales")
    assert is_valid_collection_name("销售 2024")
    assert not is_valid_collection_name("")
    assert not is_valid_collection_name("   ")
    assert not is_valid_collection_name("../evil")
    assert not is_valid_collection_name("a/b")
    assert not is_valid_collection_name("a\\b")
    assert not is_valid_collection_name("..")


def test_skips_preamble_and_finds_real_header(tmp_path):
    csv = tmp_path / "sales.csv"
    csv.write_text(
        "2024年Q4销售汇总\n\n数据来源:ERP,更新:2024-12-31\n"
        "订单号,城市,销售额,数量\n1001,北京,1200.5,3\n1002,上海,880,5\n",
        encoding="utf-8",
    )
    res = import_workbooks([csv], dest_dir=tmp_path / "imports")
    sheet = res.manifest.workbooks[0].sheets[0]
    assert sheet.header_row == 3                       # title + blank + meta skipped
    assert [c.name for c in sheet.columns] == ["订单号", "城市", "销售额", "数量"]
    assert sheet.row_count == 2


def test_plain_header_row_zero_no_regression(tmp_path):
    csv = tmp_path / "plain.csv"
    csv.write_text("a,b\n1,2\n3,4\n", encoding="utf-8")
    res = import_workbooks([csv], dest_dir=tmp_path / "imports")
    assert res.manifest.workbooks[0].sheets[0].header_row == 0


def test_all_text_table_falls_back_to_first_row(tmp_path):
    csv = tmp_path / "names.csv"
    csv.write_text("name,city\nAda,BJ\nBob,SH\n", encoding="utf-8")
    res = import_workbooks([csv], dest_dir=tmp_path / "imports")
    sheet = res.manifest.workbooks[0].sheets[0]
    assert sheet.header_row == 0
    assert [c.name for c in sheet.columns] == ["name", "city"]


def test_user_selected_header_row_matches_columns_below(tmp_path):
    from dbaide.ingest import ImportSpec

    csv = tmp_path / "report.csv"
    # auto-detect would pick row 1; force the user's choice of row 0 instead.
    csv.write_text("region,q1,q2\ncode,jan,feb\nN,10,20\nS,30,40\n", encoding="utf-8")
    res = import_workbooks(
        [ImportSpec(csv, name="report", header_anchors={"report": (0, 0)})], dest_dir=tmp_path / "imports"
    )
    sheet = res.manifest.workbooks[0].sheets[0]
    assert sheet.header_row == 0
    assert [c.name for c in sheet.columns] == ["region", "q1", "q2"]
    assert sheet.row_count == 3                         # the row the auto-detector saw as header is now data


def test_user_selected_start_column_excludes_left_columns(tmp_path):
    from dbaide.ingest import ImportSpec

    csv = tmp_path / "x.csv"
    # a junk column on the left the user wants excluded; pick start column 2.
    csv.write_text("备注,,订单号,数量\n忽略,,1001,3\n忽略,,1002,5\n", encoding="utf-8")
    res = import_workbooks(
        [ImportSpec(csv, name="x", header_anchors={"x": (0, 2)})], dest_dir=tmp_path / "imports"
    )
    sheet = res.manifest.workbooks[0].sheets[0]
    assert [c.name for c in sheet.columns] == ["订单号", "数量"]    # 备注 + blank col dropped
    assert list(sheet.data_bbox) == [0, 2, 2, 3]


def test_auto_detect_column_span_from_header_not_stray_margin(tmp_path):
    csv = tmp_path / "s.csv"
    # header is in cols B/C (col A header blank); a stray note sits in col A of a data row.
    csv.write_text(",H1,H2\nnote,1,2\n,3,4\n", encoding="utf-8")
    res = import_workbooks([csv], dest_dir=tmp_path / "imports")
    sheet = res.manifest.workbooks[0].sheets[0]
    assert [c.name for c in sheet.columns] == ["H1", "H2"]   # not dragged left to a blank col_1
    assert list(sheet.data_bbox) == [0, 1, 2, 2]


def test_ingest_replace_refuses_non_collection_connection(tmp_path):
    from dbaide.cli import build_parser, dispatch_ingest
    from dbaide.config import ConfigManager
    from dbaide.models import ConnectionConfig

    cfg = ConfigManager(path=tmp_path / "config.toml")
    cfg.upsert_connection(ConnectionConfig(name="prod", type="mysql", host="h"))
    csv = tmp_path / "x.csv"; csv.write_text("a\n1\n", encoding="utf-8")
    args = build_parser().parse_args(["ingest", str(csv), "--conn", "prod", "--replace"])
    assert dispatch_ingest(args, cfg) == 1                    # prod isn't an Excel collection


def test_excluded_left_column_does_not_create_junk_rows(tmp_path):
    from dbaide.ingest import ImportSpec

    csv = tmp_path / "r.csv"
    # an excluded left column has stray cells in otherwise-blank rows, plus a footer note
    # far below the real table — neither should leak into the table.
    csv.write_text("note,id,name\na,1,x\nb,2,y\nz,,\n,,\nfooter,,\n", encoding="utf-8")
    res = import_workbooks(
        [ImportSpec(csv, name="r", header_anchors={"r": (0, 1)})], dest_dir=tmp_path / "imports"
    )
    sheet = res.manifest.workbooks[0].sheets[0]
    assert list(sheet.data_bbox) == [0, 1, 2, 2]       # last_row = 2, not dragged to the footer
    assert sheet.row_count == 2
    con = sqlite3.connect(res.db_path)
    try:
        assert con.execute(f'SELECT id, name FROM {_q(sheet.table)}').fetchall() == [(1, "x"), (2, "y")]
    finally:
        con.close()


def test_bad_sheet_skipped_others_imported(tmp_path):
    pytest.importorskip("openpyxl")
    from openpyxl import Workbook

    wb = Workbook()
    good = wb.active
    good.title = "good"
    good.append(["a", "b"]); good.append([1, 2])
    wb.create_sheet("blank")            # empty sheet → no table → skipped, not fatal
    path = tmp_path / "mixed.xlsx"
    wb.save(path)

    res = import_workbooks([path], dest_dir=tmp_path / "imports")
    sheets = res.manifest.workbooks[0].sheets
    assert [s.sheet_name for s in sheets] == ["good"]
    assert any("blank" in w for w in res.warnings)


def test_all_sheets_unusable_raises(tmp_path):
    pytest.importorskip("openpyxl")
    from openpyxl import Workbook

    wb = Workbook()
    wb.active.title = "empty"          # only an empty sheet
    path = tmp_path / "empty.xlsx"
    wb.save(path)
    with pytest.raises(ValueError):
        import_workbooks([path], dest_dir=tmp_path / "imports")


def test_vertical_merges_fill_groups_but_keep_real_blanks(tmp_path):
    pytest.importorskip("openpyxl")
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "区域"
    ws.append(["大区", "省", "城市", "销售额", "备注"])
    ws.append(["华北", "河北", "石家庄", 400, None])
    ws.append([None, None, "唐山", 300, None])
    ws.append([None, "河南", "郑州", 500, None])
    ws.append([None, None, "洛阳", 200, "重点"])
    ws.append(["华南", "广东", "广州", 900, None])
    ws.merge_cells("A2:A5")
    ws.merge_cells("B2:B3")
    ws.merge_cells("B4:B5")
    path = tmp_path / "region.xlsx"
    wb.save(path)

    res = import_workbooks([path], dest_dir=tmp_path / "imports")
    sheet = res.manifest.workbooks[0].sheets[0]
    assert sheet.filled_columns == [0, 1]              # 大区 / 省 filled; 备注 (no merge) not
    con = sqlite3.connect(res.db_path)
    try:
        rows = con.execute(f'SELECT 大区, 省, 城市, 备注 FROM {_q(sheet.table)}').fetchall()
    finally:
        con.close()
    assert rows == [
        ("华北", "河北", "石家庄", None),
        ("华北", "河北", "唐山", None),
        ("华北", "河南", "郑州", None),
        ("华北", "河南", "洛阳", "重点"),               # the one genuine value, not forward-filled
        ("华南", "广东", "广州", None),
    ]


def test_overwrite_add_is_atomic_on_failure(tmp_path, monkeypatch):
    """A failed multi-file overwrite-add must leave the collection exactly as it was — the
    existing table keeps its rows and no partial new table appears (regression: sqlite3
    auto-commits DDL, so a plain close-without-commit left the old table dropped)."""
    import dbaide.ingest.importer as imp
    from dbaide.ingest import ExcelCollection, ImportSpec

    v1 = tmp_path / "v1.csv"; v1.write_text("amt\n1\n2\n3\n", encoding="utf-8")
    v2 = tmp_path / "v2.csv"; v2.write_text("amt\n10\n", encoding="utf-8")
    good = tmp_path / "good.csv"; good.write_text("k\n9\n", encoding="utf-8")
    col = ExcelCollection(tmp_path / "imports" / "c")
    col.add([ImportSpec(v1, name="x")])

    orig = imp._write_table
    calls = {"n": 0}

    def boom(*a, **k):
        calls["n"] += 1
        if calls["n"] >= 2:
            raise RuntimeError("simulated mid-batch failure")
        return orig(*a, **k)

    monkeypatch.setattr(imp, "_write_table", boom)
    with pytest.raises(RuntimeError):
        col.add([ImportSpec(v2, name="x"), ImportSpec(good, name="good")], overwrite=True)
    monkeypatch.undo()

    assert [(w.name, w.sheets[0].row_count) for w in col.workbooks()] == [("x", 3)]
    con = sqlite3.connect(col.db_path)
    try:
        tables = {r[0] for r in con.execute("SELECT name FROM sqlite_master WHERE type='table'")}
        assert tables == {"x"}                                            # no partial "good"
        assert con.execute("SELECT count(*) FROM x").fetchone()[0] == 3   # old rows intact
    finally:
        con.close()


def test_collection_for_connection_detects_imports(tmp_path):
    from dbaide.ingest import ExcelCollection, collection_dir, collection_for_connection

    cfg_dir = tmp_path / "cfg"
    col = ExcelCollection(collection_dir(cfg_dir, "shop"))
    col.add([_write(tmp_path / "s.csv", "a\n1\n")])

    found = collection_for_connection(cfg_dir, col.db_path)
    assert found is not None and found.dir == col.dir
    # an ordinary sqlite file elsewhere is not a collection
    assert collection_for_connection(cfg_dir, tmp_path / "random.db") is None


def _write(path: Path, text: str) -> Path:
    path.write_text(text, encoding="utf-8")
    return path


def test_cli_ingest_registers_sqlite_connection(tmp_path):
    from dbaide.cli import build_parser, dispatch_ingest
    from dbaide.config import ConfigManager

    csv = tmp_path / "people.csv"
    csv.write_text("name,age\nAda,30\n", encoding="utf-8")
    cfg = ConfigManager(path=tmp_path / "config.toml")

    args = build_parser().parse_args(["ingest", str(csv), "--conn", "people", "--default"])
    assert dispatch_ingest(args, cfg) == 0

    conns = cfg.connections()
    assert "people" in conns
    assert conns["people"].type == "sqlite"
    assert Path(conns["people"].path).exists()

    # re-ingesting the same name without --replace is refused
    args2 = build_parser().parse_args(["ingest", str(csv), "--conn", "people"])
    assert dispatch_ingest(args2, cfg) == 1
