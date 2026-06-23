"""Read a CSV/Excel file into clean rectangular sheets.

Beyond P1 (header on row 1), this locates the real table inside a sheet: it skips
title/preamble rows above the header, fills vertically-merged grouping columns downward,
and matches the columns under the chosen header. The header anchor (row + start column) can
be detected automatically (type-coherence) or supplied by the user.

Reading is per-sheet: a sheet that fails to parse is skipped with a warning, the rest of the
workbook still imports. Out of scope (fall back to "header = first non-empty row"): multi-
level / merged *headers*, crosstab unpivot, multiple tables per sheet, transpose.
"""

from __future__ import annotations

import csv
import datetime as _dt
import io
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

_CSV_EXTS = {".csv", ".tsv", ".txt"}
_XLSX_EXTS = {".xlsx", ".xlsm"}
SUPPORTED_EXTS = _CSV_EXTS | _XLSX_EXTS

_EMPTY, _NUM, _DATE, _TEXT = "e", "n", "d", "t"
_NUM_RE = re.compile(r"[+-]?(\d+\.?\d*|\.\d+)([eE][+-]?\d+)?$")
_DETECT_ROW_CAP = 30
_BODY_WINDOW = 8
_MAX_GRID_ROWS = 200_000    # guard against openpyxl reporting an inflated sheet dimension
_MAX_GRID_COLS = 1024

Anchor = tuple[int, int]   # (header_row, start_col), 0-based


@dataclass
class RawSheet:
    name: str
    header: list[str]
    rows: list[list[Any]]
    header_row: int
    data_bbox: tuple[int, int, int, int]      # (r0, c0, r1, c1) of the located table
    filled_columns: list[int] = field(default_factory=list)


@dataclass
class RawWorkbook:
    filename: str
    sheets: list[RawSheet]
    warnings: list[str] = field(default_factory=list)   # sheets that were skipped, and why


@dataclass
class RawGrid:
    cells: list[list[Any]]
    merges: list[tuple[int, int, int, int]]
    nrows: int
    ncols: int


@dataclass
class SheetGrid:
    """Filled grid + auto-detected header anchor, for the UI preview / header picker."""
    name: str
    cells: list[list[Any]]
    auto_header_row: int
    auto_header_col: int
    filled_columns: list[int]


def read_workbook(path: Path, *, header_overrides: dict[str, Anchor] | None = None,
                  sheets: list[str] | None = None) -> RawWorkbook:
    """Read a file into shaped sheets, one sheet at a time. ``header_overrides`` maps a sheet
    name to a user-chosen ``(header_row, start_col)`` anchor (else it's auto-detected);
    ``sheets``, if given, restricts the import to those sheet names. A sheet that errors is
    skipped and noted in ``warnings``; raises only if no sheet is usable."""
    path = Path(path)
    overrides = header_overrides or {}
    selected = set(sheets) if sheets is not None else None
    out_sheets: list[RawSheet] = []
    warnings: list[str] = []
    for name, grid, err in _build_grids(path):
        if selected is not None and name not in selected:
            continue                              # user deselected this sheet
        if grid is None:
            warnings.append(f"{name}: {err or 'unreadable'}")
            continue
        try:
            filled = _fill_vertical_merges(grid)
            sheet = _shape_sheet(name, grid, filled, anchor=overrides.get(name))
            if sheet is None:
                warnings.append(f"{name}: empty / no table found")
            else:
                out_sheets.append(sheet)
        except Exception as exc:  # noqa: BLE001 - isolate one sheet's failure
            warnings.append(f"{name}: {exc}")
    if not out_sheets:
        detail = f" ({'; '.join(warnings)})" if warnings else ""
        raise ValueError(f"no usable sheet in {path.name}{detail}")
    return RawWorkbook(filename=path.name, sheets=out_sheets, warnings=warnings)


def read_sheet_grids(path: Path) -> list[SheetGrid]:
    """Merge-filled grids + auto-detected anchors per sheet — feeds the preview/header picker.
    Unreadable sheets are silently skipped (the picker can only show what parsed)."""
    out: list[SheetGrid] = []
    for name, grid, _err in _build_grids(Path(path)):
        if grid is None:
            continue
        filled = _fill_vertical_merges(grid)
        loc = locate_table(grid)
        out.append(SheetGrid(
            name=name, cells=grid.cells,
            auto_header_row=(loc[0] if loc else 0),
            auto_header_col=(loc[1] if loc else 0),
            filled_columns=filled,
        ))
    return out


# ── shaping a located table into a RawSheet ──────────────────────────────────

def _shape_sheet(name: str, grid: RawGrid, filled: list[int], *, anchor: Anchor | None) -> RawSheet | None:
    if anchor is None:
        loc = locate_table(grid)
    else:
        loc = locate_table(grid, header_row=anchor[0], header_col=anchor[1])
    if loc is None:
        return None
    h, c0, c1, last = loc
    header = [_cell_text(grid.cells[h][c]) for c in range(c0, c1 + 1)]
    rows = [
        [grid.cells[r][c] for c in range(c0, c1 + 1)]
        for r in range(h + 1, last + 1)
        if _slice_nonempty(grid.cells[r], c0, c1)
    ]
    return RawSheet(
        name=name, header=header, rows=rows, header_row=h,
        data_bbox=(h, c0, last, c1),
        filled_columns=[c for c in filled if c0 <= c <= c1],
    )


# ── header / table location ──────────────────────────────────────────────────

def locate_table(grid: RawGrid, *, header_row: int | None = None, header_col: int | None = None
                 ) -> tuple[int, int, int, int] | None:
    """Locate the table as (header_row, c0, c1, last_row). ``header_row`` / ``header_col`` pin
    the header anchor (the user's pick); unset values are detected / taken from the non-empty
    extent. Columns left of ``header_col`` are excluded. Returns None for an empty sheet."""
    cells, nrows, ncols = grid.cells, grid.nrows, grid.ncols
    nonempty = [r for r in range(nrows) if _row_nonempty(cells[r])]
    if not nonempty:
        return None
    h = max(0, min(header_row, nrows - 1)) if header_row is not None else _detect_header(cells, nonempty, ncols)
    # Columns come from the HEADER row's non-empty extent — the header labels define the table.
    # Using the union of all body rows instead would let a stray value in a margin column drag
    # the table edge out (and an entirely empty chosen row would have no columns at all).
    header_cols = [c for c in range(ncols) if _nonempty(cells[h][c])]
    if not header_cols:
        return None
    c0 = max(0, min(header_col, ncols - 1)) if header_col is not None else min(header_cols)
    right = [c for c in header_cols if c >= c0]
    if not right:
        return None
    c1 = max(right)
    # Emptiness is judged within the table's column span, so a value in an excluded left
    # column (c < c0) neither emits a junk row nor drags last_row down to a stray footer.
    body = [r for r in nonempty if r > h]
    in_span = [r for r in body if _slice_nonempty(cells[r], c0, c1)]
    return h, c0, c1, (in_span[-1] if in_span else h)


def _detect_header(cells: list[list[Any]], nonempty: list[int], ncols: int) -> int:
    for h in nonempty[:_DETECT_ROW_CAP]:
        below = [r for r in nonempty if r > h][:_BODY_WINDOW]
        if below and _is_header_row(cells, h, below, ncols):
            return h
    return nonempty[0]


def _is_header_row(cells: list[list[Any]], h: int, below: list[int], ncols: int) -> bool:
    transition = False
    for c in range(ncols):
        if _ctype(cells[h][c]) != _TEXT:
            continue
        col = [_ctype(cells[r][c]) for r in below if _nonempty(cells[r][c])]
        if col and col[0] in (_NUM, _DATE) and _ratio(col, (_NUM, _DATE)) >= 0.7:
            transition = True
            break
    if not transition:
        return False
    htypes = [_ctype(cells[h][c]) for c in range(ncols) if _nonempty(cells[h][c])]
    return bool(htypes) and _ratio(htypes, (_TEXT,)) >= 0.5


def _ratio(types: list[str], wanted: tuple[str, ...]) -> float:
    return sum(t in wanted for t in types) / len(types)


def _ctype(v: Any) -> str:
    if v is None or (isinstance(v, str) and v.strip() == ""):
        return _EMPTY
    if isinstance(v, bool):
        return _NUM
    if isinstance(v, (int, float)):
        return _NUM
    if isinstance(v, (_dt.datetime, _dt.date)):
        return _DATE
    return _NUM if _NUM_RE.match(str(v).strip()) else _TEXT


# ── merged-cell fill (vertical grouping columns only) ────────────────────────

def _fill_vertical_merges(grid: RawGrid) -> list[int]:
    filled: set[int] = set()
    for (r0, c0, r1, c1) in grid.merges:
        if c0 != c1 or r1 <= r0:
            continue
        if c0 >= grid.ncols or r0 >= grid.nrows:     # merge beyond the materialized grid
            continue
        r1 = min(r1, grid.nrows - 1)                  # clamp (grid may be capped/trimmed)
        value = grid.cells[r0][c0]
        if value is None or (isinstance(value, str) and value.strip() == ""):
            continue
        for r in range(r0, r1 + 1):
            grid.cells[r][c0] = value
        filled.add(c0)
    return sorted(filled)


# ── building the raw grid (per-sheet tolerant) ───────────────────────────────

def _build_grids(path: Path) -> list[tuple[str, RawGrid | None, str]]:
    """Return (sheet_name, grid_or_None, error) per sheet. A per-sheet read error yields a
    None grid + message instead of aborting the file. Unsupported types / unopenable files
    still raise (nothing to be resilient about)."""
    ext = path.suffix.lower()
    if ext in _CSV_EXTS:
        try:
            name, grid = _csv_grid(path)
            return [(name, grid, "")]
        except Exception as exc:  # noqa: BLE001
            return [(path.stem, None, str(exc))]
    if ext in _XLSX_EXTS:
        return _xlsx_grids(path)
    raise ValueError(f"unsupported file type {ext!r}; supported: {', '.join(sorted(SUPPORTED_EXTS))}")


def _csv_grid(path: Path) -> tuple[str, RawGrid]:
    text = _decode(path.read_bytes())
    sample = text[:8192]
    try:
        delimiter = csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
    except csv.Error:
        delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    rows = list(csv.reader(io.StringIO(text), delimiter=delimiter))
    ncols = max((len(r) for r in rows), default=0)
    cells = [[(r[c] if c < len(r) else None) for c in range(ncols)] for r in rows]
    return path.stem, RawGrid(cells=cells, merges=[], nrows=len(cells), ncols=ncols)


def _xlsx_grids(path: Path) -> list[tuple[str, RawGrid | None, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - environment-dependent
        raise RuntimeError(
            "reading .xlsx/.xlsm requires openpyxl — install it with: pip install openpyxl"
        ) from exc
    # Not read_only: read_only worksheets don't expose merged-cell ranges, which we need.
    wb = load_workbook(path, read_only=False, data_only=True)
    try:
        out: list[tuple[str, RawGrid | None, str]] = []
        for ws in wb.worksheets:
            if ws.sheet_state != "visible":
                continue
            try:
                # iter_rows (capped) instead of a max_row×max_col cell loop: a stray formatted
                # cell can inflate the reported dimension to the whole sheet (1M+ rows).
                raw: list[list[Any]] = []
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i >= _MAX_GRID_ROWS:
                        break
                    raw.append(list(row[:_MAX_GRID_COLS]))
                while raw and not _row_nonempty(raw[-1]):
                    raw.pop()
                ncols = max((len(r) for r in raw), default=0)
                cells = [r + [None] * (ncols - len(r)) for r in raw]
                merges = [
                    (rng.min_row - 1, rng.min_col - 1, rng.max_row - 1, rng.max_col - 1)
                    for rng in ws.merged_cells.ranges
                ]
                out.append((ws.title, RawGrid(cells=cells, merges=merges, nrows=len(cells), ncols=ncols), ""))
            except Exception as exc:  # noqa: BLE001 - isolate one sheet
                out.append((ws.title, None, str(exc)))
        return out
    finally:
        wb.close()


def _decode(raw: bytes) -> str:
    for enc in ("utf-8-sig", "utf-8", "gb18030", "latin-1"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def _row_nonempty(row: list[Any]) -> bool:
    return any(_nonempty(c) for c in (row or []))


def _slice_nonempty(row: list[Any], c0: int, c1: int) -> bool:
    return any(_nonempty(row[c]) for c in range(c0, min(c1 + 1, len(row))))


def _nonempty(cell: Any) -> bool:
    return cell is not None and str(cell).strip() != ""


def _cell_text(cell: Any) -> str:
    return "" if cell is None else str(cell).strip()
